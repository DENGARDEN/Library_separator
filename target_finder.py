## Adapted: https://www.geeksforgeeks.org/how-to-iterate-through-excel-rows-in-python/
## https://stackoverflow.com/questions/22613272/how-to-access-the-real-value-of-a-cell-using-the-openpyxl-module-for-python
## User interaction should be implemented
## https://www.nature.com/articles/nmeth.4104
# Will be integrated into Indel searcher

import os

# import module
import openpyxl

# load excel with its path
wrkbk = openpyxl.load_workbook("AandT.xlsx", data_only=True, read_only=True)
sh = wrkbk.active
PATH = "./References"

if not os.path.exists(PATH):
    os.makedirs(PATH)

INDEX_COL = 4
WIDE_TARGET_COL = 9
RP_BINDING_COL = 10
TYPE_COL = 2

target_path = os.path.join(PATH, "Target_region.txt")

with open(target_path, 'w') as target:
    # iterate through excel and display data
    for i in range(2, sh.max_row + 1):
        print("\n")
        print("Row ", i, " data :")
        wide_target_sequence = sh.cell(row=i, column=WIDE_TARGET_COL)
        rev_homology_sequence = sh.cell(row=i, column=RP_BINDING_COL)

        chunk = wide_target_sequence.value + rev_homology_sequence.value
        print(chunk, end=" ")

        type = sh.cell(row=i, column=TYPE_COL).value
        category = sh.cell(row=i, column=INDEX_COL).value

        if type == "AsCas12f" and category == "AsFullMatch":
            UP_CONTEXT = 34
            PAM_LENGTH = 4
            WEDGE_POS = 20
            WINDOW_SIZE = 10
        elif type == "AsCas12f":
            UP_CONTEXT = 4
            PAM_LENGTH = 4
            WEDGE_POS = 20
            WINDOW_SIZE = 10

        elif type == "TnpB" and category == "TFullMatch":
            UP_CONTEXT = 30
            PAM_LENGTH = 5
            WEDGE_POS = 18
            WINDOW_SIZE = 10

        elif type == "TnpB":
            UP_CONTEXT = 4
            PAM_LENGTH = 5
            WEDGE_POS = 18
            WINDOW_SIZE = 10

        # Save as files
        target.write(f"{chunk[UP_CONTEXT: UP_CONTEXT + PAM_LENGTH + WEDGE_POS + WINDOW_SIZE]}\n")
