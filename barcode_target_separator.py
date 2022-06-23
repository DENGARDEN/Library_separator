## Adapted: https://www.geeksforgeeks.org/how-to-iterate-through-excel-rows-in-python/
## https://stackoverflow.com/questions/22613272/how-to-access-the-real-value-of-a-cell-using-the-openpyxl-module-for-python
## User interaction should be implemented
## https://www.nature.com/articles/nmeth.4104
# Will be integrated into Indel searcher

# import module
import openpyxl
import os

# load excel with its path
wrkbk = openpyxl.load_workbook("1st order sheet.xlsx", data_only = True)

sh = wrkbk.active


# Parameters
ref_col = 2
length_col = 3


FWD_HOMOLOGY_LENGTH = 27
GUIDE_LENGTH = 20
POLY_T_LENGTH = 7
BARCODE_LENGTH = 24
PAM_TYPE = "Cas12a"
PAM_LENGTH =  4 if PAM_TYPE == "Cas12a" else 3
TARGET_LENGTH = 26
REV_HOMOLOGY_LENGTH = 22
WINDOW_SIZE = 5
PATH = "./References"

WEDGE_POS = 22

# Preliminary calculations
s_barcode = FWD_HOMOLOGY_LENGTH+ GUIDE_LENGTH+ POLY_T_LENGTH
e_barcode = s_barcode + BARCODE_LENGTH

# Starting from PAM; TTTV
s_guide = e_barcode
e_guide = s_guide+PAM_LENGTH + GUIDE_LENGTH 



if not os.path.exists(PATH):
	os.makedirs(PATH)

barcode_path = os.path.join(PATH, "Barcode.txt")
target_path = os.path.join(PATH, "Target_region.txt")
ref_path = os.path.join(PATH, "Reference_sequence.txt")

with open(barcode_path,'w') as barcode, \
	open(target_path,'w' ) as target, \
	open(ref_path,'w') as ref:
	# iterate through excel and display data
	for i in range(1, sh.max_row+1):
		assert (FWD_HOMOLOGY_LENGTH+GUIDE_LENGTH+POLY_T_LENGTH +BARCODE_LENGTH+ PAM_LENGTH +TARGET_LENGTH + REV_HOMOLOGY_LENGTH ) \
			== int(sh.cell(row=i , column =length_col).value), "Parameters are not matched with the source data"
		print("\n")
		print("Row ", i, " data :")
		cell_obj = sh.cell(row=i, column=ref_col)
		print(cell_obj.value, end=" ")

		# Save as files
		barcode.write(f"{cell_obj.value[s_barcode: e_barcode]}\n")
		target.write(f"{cell_obj.value[s_guide: s_guide+PAM_LENGTH+ WEDGE_POS+WINDOW_SIZE]}\n")
		ref.write(f"{cell_obj.value}\n")


